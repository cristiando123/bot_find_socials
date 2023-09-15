from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
from selenium.webdriver.common.by import By
import openpyxl
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog
import undetected_chromedriver as uc
import re
import os




def clean_input(input_str):
    # Loại bỏ các ký tự đặc biệt, chỉ giữ lại ký tự chữ cái và dấu cách
    cleaned_str = re.sub(r'[^a-zA-Z\s]', '', input_str)
    return cleaned_str



def select_file():
    file_path = filedialog.askopenfilename(filetypes=(("All files", "*"),))
    file_entry.delete(0, tk.END)
    file_entry.insert(0, file_path)

def search_twitter_profile():
    try:
        # Load the input Excel file
        input_file_path = file_entry.get()

        # Open the input workbook and get the first sheet
        input_workbook = openpyxl.load_workbook(input_file_path)
        input_sheet = input_workbook.active

        # Create a new workbook for the output file
        output_workbook = Workbook()
        output_sheet = output_workbook.active

        output_sheet.append(["CEO Name", "Company Name", "Keywords", "Results"])

        # Loop through each row in the input sheet
        for row in input_sheet.iter_rows(min_row=2, values_only=True):
            if any(cell_value is not None and cell_value != "" for cell_value in row):
                ceo_name = row[0] if row[0] else ""
                company_name = row[1] if row[1] else ""
                keywords = row[2] if row[2] else ""

                query = f"{company_name} {ceo_name} {keywords} Twitter account"
                query = query.replace(" ", "%20")

                search_url = f"https://www.google.com/search?q={query}"

                # Use regular Chrome WebDriver
                driver = webdriver.Chrome()
                driver.get(search_url)
                driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.END)
                page_source = driver.page_source
                driver.quit()

                soup = BeautifulSoup(page_source, "html.parser")
                search_results = soup.find_all('a', href=True)  # Find all anchor tags with href attribute

                for row in input_sheet.iter_rows(min_row=2, values_only=True):
    # ... (Các phần mã khác)

                    twitter_links = []  # Danh sách các liên kết Twitter cho từng dòng

                    for result in search_results:
                        link = result['href']
                        if 'twitter.com' in link:
                            twitter_links.append(link)

                    # Thêm danh sách liên kết Twitter vào dòng đầu ra
                    if twitter_links:
                        for twitter_link in twitter_links:
                            output_sheet.append([ceo_name, company_name, keywords, twitter_link])
                    else:
        # Nếu không có liên kết Twitter, thêm một dòng với thông báo "No Twitter links found."
                        output_sheet.append([ceo_name, company_name, keywords, "No Twitter links found."])
        # Save the output workbook with the Twitter profiles for each row
        output_file_path = "output_twitter_file.xlsx"
        output_workbook.save(output_file_path)
        status_label.config(text=f"Output saved to: {output_file_path}")



    except Exception as e:
        status_label.config(text="Error occurred while processing the Excel file.")

def search_facebook_profile():
    try:
        # Load the input Excel file
        input_file_path = file_entry.get()

        # Open the input workbook and get the first sheet
        input_workbook = openpyxl.load_workbook(input_file_path)
        input_sheet = input_workbook.active

        # Create a new workbook for the output file
        output_workbook = Workbook()
        output_sheet = output_workbook.active

        output_sheet.append(["CEO Name","Company Name" ,"Keywords" , "Results"])

        # Loop through each row in the input sheet
        for row in input_sheet.iter_rows(min_row=2, values_only=True):
            if any(cell_value is not None and cell_value != "" for cell_value in row): 
                ceo_name = row[0] if row[0] else ""
                company_name = row[1] if row[1] else ""
                keywords = row[2] if row[2] else ""
                
                #if company_name and company_name:
                query = f"{company_name} {ceo_name} {keywords} Facebook account"
                query = query.replace(" ", "%20")

                search_url = f"https://www.google.com/search?q={query}"

                driver = webdriver.Chrome()
                driver.get(search_url)
                driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.END)
                page_source = driver.page_source
                driver.quit()

                soup = BeautifulSoup(page_source, "html.parser")
                search_results = soup.find_all('a', href=True)  # Find all anchor tags with href attribute

                
                for row in input_sheet.iter_rows(min_row=2, values_only=True):
    # ... (Các phần mã khác)

                    facebook_links = []  # Danh sách các liên kết Twitter cho từng dòng

                    for result in search_results:
                        link = result['href']
                        if 'facebook.com' in link:
                            facebook_links.append(link)

                    # Thêm danh sách liên kết Twitter vào dòng đầu ra
                    if facebook_links:
                        for twitter_link in facebook_links:
                            output_sheet.append([ceo_name, company_name, keywords, facebook_links])
                    else:
        # Nếu không có liên kết Twitter, thêm một dòng với thông báo "No Twitter links found."
                        output_sheet.append([ceo_name, company_name, keywords, "No facebook links found."])

        # Save the output workbook with the Twitter profiles for each row
        output_file_path = "output_facebook_file.xlsx"
        output_workbook.save(output_file_path)
        status_label.config(text=f"Output saved to: {output_file_path}")

    except Exception as e:
        status_label.config(text="Error occurred while processing the Excel file.")

def search_linkedin_profile():
    try:
        # Load the input Excel file
        input_file_path = file_entry.get()

        # Open the input workbook and get the first sheet
        input_workbook = openpyxl.load_workbook(input_file_path)
        input_sheet = input_workbook.active

        # Create a new workbook for the output file
        output_workbook = Workbook()
        output_sheet = output_workbook.active

        output_sheet.append(["CEO Name","Company Name" ,"Keywords" , "Results"])

        # Loop through each row in the input sheet
        for row in input_sheet.iter_rows(min_row=2, values_only=True):
            if any(cell_value is not None and cell_value != "" for cell_value in row): 
                ceo_name = row[0] if row[0] else ""
                company_name = row[1] if row[1] else ""
                keywords = row[2] if row[2] else ""
                
                #if company_name and company_name:
                query = f"{company_name} {ceo_name} {keywords} LinkedIn account"
                query = query.replace(" ", "%20")

                search_url = f"https://www.google.com/search?q={query}"

                driver = webdriver.Chrome()
                driver.get(search_url)
                driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.END)
                page_source = driver.page_source
                driver.quit()

                soup = BeautifulSoup(page_source, "html.parser")
                search_results = soup.find_all('a', href=True)  # Find all anchor tags with href attribute

                for row in input_sheet.iter_rows(min_row=2, values_only=True):
                    linkedin_links = []
                    for result in search_results:
                            link = result['href']
                            if 'linkedin.com' in link:
                                linkedin_links.append(link)

                    if linkedin_links:
                        for linkedin_links in linkedin_links:
                            output_sheet.append([ceo_name, company_name, keywords, linkedin_links])
                    else:
                        output_sheet.append([ceo_name, company_name, keywords, "No linkedin links found."])
                
                
                
                # linkedin_links = []
                # count = 0  # Counter for the number of linkedin links collected

                # for result in search_results:
                #     link = result['href']
                #     if 'linkedin.com' in link:
                #         linkedin_links.append(link)
                #         count += 1

                #     if count == 3:  # If three linkedin links are found, break the loop
                #         break

                # if linkedin_links:
                #     output_sheet.append([ceo_name,company_name,keywords, '\n'.join(linkedin_links)])
                # else:
                #     output_sheet.append([ceo_name,company_name,keywords, "No LinkedIn links found."])

        # Save the output workbook with the linkedin profiles for each row
        output_file_path = "output_linkedin_file.xlsx"
        output_workbook.save(output_file_path)
        status_label.config(text=f"Output saved to: {output_file_path}")

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        status_label.config(text="Error occurred while processing the Excel file.")

root = tk.Tk()
root.title("Find Social Media account")

# create the main frame
main_frame = tk.Frame(root, padx=10, pady=10)
main_frame.pack()

# create the file frame
file_frame = tk.LabelFrame(main_frame, text="Attached input file")
file_frame.pack(fill="x", padx=10, pady=10)

file_entry = tk.Entry(file_frame, width=40)
file_entry.pack(side="left", padx=10, pady=5)

file_button = tk.Button(file_frame, text="Browser", command=select_file)
file_button.pack(side="left", padx=10, pady=5)

# create the status label
status_label = tk.Label(main_frame, text="", font=("Arial", 12))
status_label.pack(pady=10)

# create the send button
send_button = tk.Button(main_frame, text="Find Twitter Account", command=search_twitter_profile)
send_button.pack(pady=5)

# create the send button
send_button = tk.Button(main_frame, text="Find Facebook Account", command=search_facebook_profile)
send_button.pack(pady=5)

# create the send button
send_button = tk.Button(main_frame, text="Find LinkedIn Account", command=search_linkedin_profile)
send_button.pack(pady=5)

root.mainloop()
